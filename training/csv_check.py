#!/usr/bin/env python
# -*- coding: utf-8 -*-

# @Time     :
# @Author   : Spike713
# @Site     :
# @File     : csv_check.py
# @Purpose  :
# @Software : PyCharm
# @Copyright:
# @Licence  :

from openpyxl.reader.excel import load_workbook

class Csvcheck():


	def __init__(self, excel_path:str):
		'''
		打开一个excel文件
		:param excel_path:
		:return:
		'''
		wb = load_workbook(excel_path)
		sheetnames = wb.get_sheet_names()
		self.ws = wb.get_sheet_by_name(sheetnames[0])

	def check_id(self):
		'''
		检查ID是否重复 是否连续  默认ID在第一列 默认ID从1开始自增
		:return:
		'''
		id_list = []
		for i in range(2, self.ws.max_row+1):
			if self.ws.cell(i, 1).value not in id_list:
				if i - 1 != self.ws.cell(i, 1).value:
					print("ID自增错误！！ 行数:{}".format(i + 1))

				id_list.append(self.ws.cell(i, 1).value)
				print(id_list)
			else:
				print("ID重复！！ 行数:{}".format(i))
"""
	def check_lua():
		'''
		检查lua数据列 是否存在中文标点符号 是否存在不连续的{}
		:return:
		'''
		cn_str = ["，", "。", "；", "："]
        	for i in range(2, ws.max_row + 1):
            		#for j in range(2, ws.max_column + 1):
                		lua_str = str(ws.cell(i, 1).value)
                		for k in cn_str:
                    			if k in lua_str:
                        			print("检查到中文字符！！ 行数:{} 列数:{}".format(i, j))
                		if lua_str.count("{") != lua_str.count("}"):
                    			print("{{和}}的个数不相等！！ 行数:{} 列数:{}".format(i, j))
"""
if __name__ == '__main__':
	s = Csvcheck('/Users/Administrator/Desktop/快速生成奖励配置/reward.xlsx')
	s.check_id()
#	s.check_lua()
